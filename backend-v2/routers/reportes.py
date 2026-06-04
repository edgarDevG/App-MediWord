# ═══════════════════════════════════════════════════════════════
# reportes.py — MediWord HSM — Exportaciones multi-formato
# Formatos soportados: XLSX (estilizado), CSV (UTF-8 BOM), PDF (branding)
# ═══════════════════════════════════════════════════════════════
import csv
import io
from datetime import date, timedelta
from typing import Optional

import pandas as pd
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, inch
from reportlab.platypus import (
    HRFlowable,
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import func as sa_func
from sqlalchemy.orm import Session, joinedload

from database import get_db
from model import Medico, MedicoAccesos, MedicoContacto, MedicoContratacion, MedicoNormativos
from routers.auth import get_current_user

router = APIRouter(tags=["Reportes"])

# ── Constantes de estilo institucional ────────────────────────
BRAND_NAVY      = "#0A2540"
BRAND_JADE      = "#0A7E6E"
BRAND_BLUE      = "#1A3C6E"
BRAND_TEAL      = "#0A5C99"
BRAND_BG_ALT    = "#F8FAFC"
BRAND_BORDER    = "#E2E8F0"
BRAND_TEXT_DARK = "#0F1923"
BRAND_TEXT_MUTED = "#64748B"
BRAND_RED       = "#DC2626"
BRAND_AMBER     = "#D97706"

CAMPOS_FECHA_NORMATIVOS = [
    ("bls_fecha_venc",         "BLS"),
    ("acls_fecha_venc",        "ACLS"),
    ("pals_fecha_venc",        "PALS"),
    ("nals_fecha_venc",        "NALS"),
    ("violencia_sexual_fecha", "Violencia Sexual"),
    ("ataques_quimicos_fecha", "Agentes Químicos"),
    ("dengue_fecha",           "Dengue"),
    ("sedacion_fecha",         "Sedación"),
    ("radioproteccion_fecha",  "Radioprotección"),
    ("manejo_dolor_fecha",     "Manejo del Dolor"),
    ("iamii_fecha",            "IAMII"),
    ("gestion_duelo_fecha",    "Gestión del Duelo"),
]


# ═══════════════════════════════════════════════════════════════
# HELPERS DE DATOS
# ═══════════════════════════════════════════════════════════════

def _build_alertas(db: Session):
    hoy = date.today()
    alertas = []

    normativos = (
        db.query(MedicoNormativos)
        .options(joinedload(MedicoNormativos.medico).joinedload(Medico.datos_hv))
        .all()
    )

    for n in normativos:
        medico = n.medico
        if medico is None:
            continue
        tipo_doc = medico.datos_hv.tipo_documento if medico.datos_hv else ""
        for campo, nombre_doc in CAMPOS_FECHA_NORMATIVOS:
            fecha = getattr(n, campo, None)
            if fecha is None:
                continue
            dias_restantes = (fecha - hoy).days
            estado = "Vencido" if dias_restantes < 0 else "Por vencer"
            alertas.append({
                "Documento": medico.documento_identidad,
                "Nombre": medico.nombre_medico,
                "Tipo Documento": tipo_doc,
                "Fecha Vencimiento": fecha.isoformat(),
                "Días Restantes": dias_restantes,
                "Estado": estado,
            })

    accesos = (
        db.query(MedicoAccesos)
        .options(joinedload(MedicoAccesos.medico).joinedload(Medico.datos_hv))
        .filter(MedicoAccesos.fecha_venc_poliza.isnot(None))
        .all()
    )

    for a in accesos:
        medico = a.medico
        if medico is None or a.fecha_venc_poliza is None:
            continue
        tipo_doc = medico.datos_hv.tipo_documento if medico.datos_hv else ""
        dias_restantes = (a.fecha_venc_poliza - hoy).days
        estado = "Vencido" if dias_restantes < 0 else "Por vencer"
        alertas.append({
            "Documento": medico.documento_identidad,
            "Nombre": medico.nombre_medico,
            "Tipo Documento": tipo_doc,
            "Fecha Vencimiento": a.fecha_venc_poliza.isoformat(),
            "Días Restantes": dias_restantes,
            "Estado": estado,
        })

    alertas.sort(key=lambda x: x["Días Restantes"])
    return alertas


# ═══════════════════════════════════════════════════════════════
# HELPERS DE EXPORTACIÓN — CSV
# ═══════════════════════════════════════════════════════════════

def _export_csv(df: pd.DataFrame, filename: str) -> StreamingResponse:
    """Genera una respuesta CSV con UTF-8 BOM para compatibilidad con Excel."""
    output = io.StringIO()
    # BOM para que Excel reconozca UTF-8 automáticamente
    output.write('\ufeff')
    df.to_csv(output, index=False, quoting=csv.QUOTE_MINIMAL)
    content = output.getvalue().encode("utf-8")
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        iter([content]),
        headers=headers,
        media_type="text/csv; charset=utf-8",
    )


# ═══════════════════════════════════════════════════════════════
# HELPERS DE EXPORTACIÓN — XLSX ESTILIZADO
# ═══════════════════════════════════════════════════════════════

def _styled_xlsx(sheets: dict[str, pd.DataFrame], filename: str) -> StreamingResponse:
    """
    Genera un XLSX con formato institucional estándar.
    sheets: dict de {nombre_hoja: DataFrame}
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            ws = writer.sheets[sheet_name]

            # ── Estilos ──
            header_fill = PatternFill("solid", fgColor="0A2540")
            header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            alt_fill_1 = PatternFill("solid", fgColor="FFFFFF")
            alt_fill_2 = PatternFill("solid", fgColor="F8FAFC")
            data_font = Font(name="Calibri", size=10, color="0F1923")
            data_align = Alignment(vertical="center", wrap_text=False)

            thin_border = Border(
                left=Side(style="thin", color="E2E8F0"),
                right=Side(style="thin", color="E2E8F0"),
                top=Side(style="thin", color="E2E8F0"),
                bottom=Side(style="thin", color="E2E8F0"),
            )

            # ── Aplicar header ──
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = thin_border
            ws.row_dimensions[1].height = 32

            # ── Aplicar datos con filas alternas ──
            for row_idx in range(2, ws.max_row + 1):
                fill = alt_fill_1 if row_idx % 2 == 0 else alt_fill_2
                for cell in ws[row_idx]:
                    cell.fill = fill
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = thin_border

            # ── Auto-width columnas ──
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                # Mínimo 10, máximo 50
                adjusted_width = min(max(max_length + 3, 10), 50)
                ws.column_dimensions[col_letter].width = adjusted_width

    output.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        headers=headers,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ═══════════════════════════════════════════════════════════════
# HELPERS DE EXPORTACIÓN — PDF ESTANDARIZADO
# ═══════════════════════════════════════════════════════════════

# Colores PDF institucionales (ReportLab HexColor)
PDF_NAVY  = HexColor(BRAND_NAVY)
PDF_JADE  = HexColor(BRAND_JADE)
PDF_BLUE  = HexColor(BRAND_BLUE)
PDF_LGREY = HexColor(BRAND_BG_ALT)
PDF_MGREY = HexColor(BRAND_BORDER)
PDF_RED   = HexColor(BRAND_RED)
PDF_AMBER = HexColor(BRAND_AMBER)


def _pdf_header_footer(canvas, doc):
    """
    Dibuja header y footer institucional en cada página del PDF.
    Header: "MEDIWORD HSM" con línea jade
    Footer: Fecha + Página n/N
    """
    canvas.saveState()
    width, height = doc.pagesize

    # ── HEADER ──
    # Línea superior jade
    canvas.setStrokeColor(HexColor(BRAND_JADE))
    canvas.setLineWidth(3)
    canvas.line(doc.leftMargin, height - doc.topMargin + 25,
                width - doc.rightMargin, height - doc.topMargin + 25)

    # Texto header
    canvas.setFont("Helvetica-Bold", 9)
    canvas.setFillColor(HexColor(BRAND_NAVY))
    canvas.drawString(doc.leftMargin, height - doc.topMargin + 32,
                      "MEDIWORD HSM")

    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(HexColor(BRAND_TEXT_MUTED))
    canvas.drawRightString(width - doc.rightMargin, height - doc.topMargin + 32,
                           "Hospital Serena del Mar · Dirección Médica")

    # ── FOOTER ──
    # Línea inferior
    canvas.setStrokeColor(HexColor(BRAND_BORDER))
    canvas.setLineWidth(0.5)
    canvas.line(doc.leftMargin, doc.bottomMargin - 10,
                width - doc.rightMargin, doc.bottomMargin - 10)

    # Fecha de generación
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(HexColor(BRAND_TEXT_MUTED))
    canvas.drawString(doc.leftMargin, doc.bottomMargin - 22,
                      f"Generado: {date.today().strftime('%d/%m/%Y %H:%M')}")

    # Número de página
    page_num = canvas.getPageNumber()
    canvas.drawRightString(width - doc.rightMargin, doc.bottomMargin - 22,
                           f"Página {page_num}")

    canvas.restoreState()


def _pdf_cell(text, bold=False, size=8, color=BRAND_TEXT_DARK, align="LEFT"):
    """Crea un Paragraph estilizado para celdas de tabla PDF."""
    al = TA_CENTER if align == "CENTER" else (TA_RIGHT if align == "RIGHT" else TA_LEFT)
    hex_color = color if color.startswith("#") else f"#{color}"
    st = ParagraphStyle(
        "cp", fontName="Helvetica-Bold" if bold else "Helvetica",
        fontSize=size, textColor=HexColor(hex_color), leading=10, alignment=al,
    )
    return Paragraph(str(text), st)


def _pdf_section_title(text, size=12, color=None):
    """Crea un título de sección para PDF."""
    c = color or PDF_BLUE
    st = ParagraphStyle(
        "hdr", fontName="Helvetica-Bold", fontSize=size,
        textColor=c, spaceAfter=2,
    )
    return Paragraph(text, st)


def _pdf_body_text(text, size=9):
    """Crea un párrafo de texto normal para PDF."""
    st = ParagraphStyle(
        "body", fontName="Helvetica", fontSize=size,
        textColor=HexColor("#334155"), leading=12,
    )
    return Paragraph(text, st)


def _pdf_standard_table_style(header_bg=None):
    """Retorna el TableStyle institucional estándar."""
    bg = header_bg or PDF_NAVY
    return TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), bg),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 9),
        ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
        ("GRID",          (0, 0), (-1, -1), 0.4, PDF_MGREY),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PDF_LGREY]),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("LINEABOVE",     (0, 0), (-1, 0), 1.5, bg),
        ("LINEBELOW",     (0, -1), (-1, -1), 0.8, PDF_MGREY),
    ])


def _build_tabular_pdf(title: str, subtitle: str, headers: list, rows: list,
                        col_widths: list, filename: str,
                        page_size=landscape(A4)) -> StreamingResponse:
    """
    Genera un PDF tabular estándar con branding institucional.
    Usado para reportes que son básicamente tablas de datos.
    """
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=page_size,
        rightMargin=1.8 * cm, leftMargin=1.8 * cm,
        topMargin=2.2 * cm, bottomMargin=2.2 * cm,
    )

    story = []

    # Título
    story.append(_pdf_section_title(title, size=16, color=PDF_NAVY))
    story.append(_pdf_body_text(
        f"Hospital Serena del Mar  ·  Generado: {date.today().strftime('%d de %B de %Y').title()}",
        size=9
    ))
    story.append(Spacer(1, 0.3 * cm))
    story.append(HRFlowable(width="100%", thickness=2, color=PDF_JADE, spaceAfter=12))

    if subtitle:
        story.append(_pdf_body_text(subtitle, size=9))
        story.append(Spacer(1, 0.3 * cm))

    if rows:
        # Header row
        header_row = [
            _pdf_cell(h, bold=True, size=8, color="FFFFFF", align="CENTER")
            for h in headers
        ]
        table_data = [header_row]

        for row in rows:
            table_data.append([_pdf_cell(str(v), size=8) for v in row])

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(_pdf_standard_table_style())
        story.append(table)
        story.append(Spacer(1, 0.3 * cm))
        story.append(_pdf_body_text(f"Total: {len(rows)} registro(s)", size=8))
    else:
        story.append(_pdf_body_text("No hay datos disponibles para este reporte.", size=10))

    doc.build(story, onFirstPage=_pdf_header_footer, onLaterPages=_pdf_header_footer)
    output.seek(0)

    resp_headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        iter([output.getvalue()]),
        headers=resp_headers,
        media_type="application/pdf",
    )


# ═══════════════════════════════════════════════════════════════
# ENDPOINT 1 — EXPORTACIÓN COMPLETA
# ═══════════════════════════════════════════════════════════════

@router.get("/exportar")
def exportar_medicos(
    formato: str = Query("xlsx", regex="^(xlsx|csv|pdf)$"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    medicos = (
        db.query(Medico)
        .options(
            joinedload(Medico.contratacion),
            joinedload(Medico.contacto),
        )
        .all()
    )

    data_medicos = []
    for m in medicos:
        c = m.contratacion
        k = m.contacto
        data_medicos.append({
            "Documento": m.documento_identidad,
            "Nombre": m.nombre_medico,
            "Categoría": m.categoria,
            "Especialidad": m.especialidad,
            "Estado": m.estado,
            "Tipo Listado": m.tipo_listado,
            "Fecha Ingreso": str(m.fecha_ingreso) if m.fecha_ingreso else "",
            "Tipo Vinculación": c.tipo_vinculacion if c else "",
            "Fecha Venc Oferta": str(c.fecha_venc_oferta) if c and c.fecha_venc_oferta else "",
            "Correo": k.correo if k else "",
            "Celular": k.celular if k else "",
        })

    alertas = _build_alertas(db)

    df_medicos = pd.DataFrame(data_medicos)
    df_alertas = pd.DataFrame(alertas)

    if formato == "csv":
        return _export_csv(df_medicos, "reporte_medicos.csv")

    if formato == "pdf":
        headers = list(df_medicos.columns)
        rows = df_medicos.values.tolist()
        W = landscape(A4)[0] - 3.6 * cm
        n = len(headers)
        col_widths = [W / n] * n
        return _build_tabular_pdf(
            title="EXPORTACIÓN COMPLETA — CUERPO MÉDICO",
            subtitle=f"Listado de {len(rows)} médicos registrados en el sistema",
            headers=headers,
            rows=rows,
            col_widths=col_widths,
            filename="reporte_medicos.pdf",
        )

    # xlsx (default)
    return _styled_xlsx(
        {"Médicos": df_medicos, "Alertas Vencimiento": df_alertas},
        "reporte_medicos.xlsx",
    )


# ═══════════════════════════════════════════════════════════════
# ENDPOINT 2 — ALERTAS DE VENCIMIENTO
# ═══════════════════════════════════════════════════════════════

@router.get("/exportar-alertas")
def exportar_alertas(
    formato: str = Query("xlsx", regex="^(xlsx|csv|pdf)$"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    alertas = [a for a in _build_alertas(db) if a["Días Restantes"] <= 30]
    df = pd.DataFrame(alertas)

    if formato == "csv":
        return _export_csv(df, "alertas_vencimiento.csv")

    if formato == "pdf":
        if alertas:
            headers = list(df.columns)
            rows = df.values.tolist()
        else:
            headers = ["Documento", "Nombre", "Tipo Documento",
                       "Fecha Vencimiento", "Días Restantes", "Estado"]
            rows = []
        W = landscape(A4)[0] - 3.6 * cm
        n = len(headers)
        col_widths = [W / n] * n
        return _build_tabular_pdf(
            title="ALERTAS DE VENCIMIENTO",
            subtitle="Documentos vencidos o próximos a vencer en los próximos 30 días",
            headers=headers,
            rows=rows,
            col_widths=col_widths,
            filename="alertas_vencimiento.pdf",
        )

    # xlsx (default)
    return _styled_xlsx({"Alertas": df}, "alertas_vencimiento.xlsx")


# ═══════════════════════════════════════════════════════════════
# ENDPOINT 3 — DIRECTORIO MÉDICO
# ═══════════════════════════════════════════════════════════════

@router.get("/directorio-medico")
def exportar_directorio_medico(
    formato: str = Query("xlsx", regex="^(xlsx|csv|pdf)$"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    from model import Departamento, MedicoDatosHV, Seccion

    medicos = (
        db.query(Medico)
        .options(
            joinedload(Medico.contacto),
            joinedload(Medico.datos_hv),
        )
        .filter(Medico.estado == "ACTIVO")
        .order_by(Medico.categoria, Medico.nombre_medico)
        .all()
    )

    # Cachear departamentos y secciones
    dptos = {d.id: d.nombre for d in db.query(Departamento).all()}
    secc = {s.id: s.nombre for s in db.query(Seccion).all()}

    rows = []
    for m in medicos:
        hv = m.datos_hv
        k = m.contacto

        # Idiomas: consolidar en un solo campo
        idiomas_arr = []
        if k and k.idiomas:
            raw = k.idiomas
            if isinstance(raw, list):
                idiomas_arr = [str(i) for i in raw if i]
            else:
                idiomas_arr = [str(raw)]
        if k and k.otro_idioma:
            idiomas_arr.append(k.otro_idioma)
        idiomas_str = ", ".join(idiomas_arr) if idiomas_arr else "—"

        tipo_listado_label = "HSM" if (m.tipo_listado or "") == "cuerpo_medico" else "FSFB"

        rows.append({
            "Tipo":                   tipo_listado_label,
            "Categoría":              m.categoria             or "—",
            "Nombre":                 m.nombre_medico         or "—",
            "Especialidad":           m.especialidad          or "—",
            "Dpto. Coordinación":     dptos.get(m.dept_coordinacion_id,     "—"),
            "Dpto. Dirección Médica": dptos.get(m.dept_direccion_medica_id, "—"),
            "Sección":                secc.get(m.seccion_id,  "—"),
            "Tipo Documento":         (hv.tipo_documento      if hv else None) or "—",
            "Número de Cédula":       m.documento_identidad   or "—",
            "Fecha Nacimiento":       hv.fecha_nacimiento.isoformat() if (hv and hv.fecha_nacimiento) else "—",
            "Correo":                 (k.correo               if k else None) or "—",
            "Teléfono":               (k.celular              if k else None) or (k.telefono if k else None) or "—",
            "Idiomas":                idiomas_str,
            "Lenguaje de Señas":      ("SÍ" if (k and k.maneja_lengua_senas) else "NO") if k else "—",
        })

    df = pd.DataFrame(rows)

    if formato == "csv":
        return _export_csv(df, "directorio_medico.csv")

    if formato == "pdf":
        if rows:
            headers = list(df.columns)
            data_rows = df.values.tolist()
        else:
            headers = ["Tipo", "Categoría", "Nombre", "Especialidad"]
            data_rows = []
        W = landscape(A4)[0] - 3.6 * cm
        # Ajustar anchos proporcionalmente
        col_widths = [W * 0.05, W * 0.07, W * 0.14, W * 0.10,
                      W * 0.10, W * 0.10, W * 0.07, W * 0.06,
                      W * 0.07, W * 0.06, W * 0.08, W * 0.05,
                      W * 0.04, W * 0.04]
        return _build_tabular_pdf(
            title="DIRECTORIO MÉDICO — PERSONAL ACTIVO",
            subtitle=f"Directorio con {len(data_rows)} médicos activos del Hospital Serena del Mar",
            headers=headers,
            rows=data_rows,
            col_widths=col_widths[:len(headers)],
            filename="directorio_medico.pdf",
        )

    # xlsx (default)
    return _styled_xlsx({"Directorio Médico": df}, "directorio_medico.xlsx")


# ═══════════════════════════════════════════════════════════════
# ENDPOINT 4 — INFORME EJECUTIVO (PDF principal + XLSX)
# ═══════════════════════════════════════════════════════════════

@router.get("/informe-ejecutivo-pdf")
def informe_ejecutivo(
    formato: str = Query("pdf", regex="^(pdf|xlsx)$"),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    hoy = date.today()

    # ── Contadores ──
    total       = db.query(sa_func.count(Medico.id)).scalar() or 0
    activos     = db.query(sa_func.count(Medico.id)).filter(Medico.estado == "ACTIVO").scalar() or 0
    finalizados = db.query(sa_func.count(Medico.id)).filter(Medico.estado == "FINALIZADO").scalar() or 0
    renuncias   = db.query(sa_func.count(Medico.id)).filter(Medico.estado == "RENUNCIA").scalar() or 0
    inactivos   = db.query(sa_func.count(Medico.id)).filter(Medico.estado == "INACTIVO").scalar() or 0
    hsm_total   = db.query(sa_func.count(Medico.id)).filter(Medico.tipo_listado == "cuerpo_medico").scalar() or 0
    fsfb_total  = db.query(sa_func.count(Medico.id)).filter(Medico.tipo_listado == "fsfb_externo").scalar() or 0

    # ── Alertas ≤30d con nombre del documento real ──
    hoy_d = date.today()
    alertas_30_enrich = []

    normativos = (
        db.query(MedicoNormativos)
        .options(joinedload(MedicoNormativos.medico))
        .all()
    )
    for n in normativos:
        med = n.medico
        if not med:
            continue
        tipo = "HSM" if med.tipo_listado == "cuerpo_medico" else "FSFB"
        for campo, nombre_doc in CAMPOS_FECHA_NORMATIVOS:
            fecha = getattr(n, campo, None)
            if not fecha:
                continue
            dias = (fecha - hoy_d).days
            if dias <= 30:
                alertas_30_enrich.append([
                    tipo, med.nombre_medico, nombre_doc,
                    fecha.strftime("%d/%m/%Y"), str(dias),
                ])

    for a in (
        db.query(MedicoAccesos)
        .options(joinedload(MedicoAccesos.medico))
        .filter(MedicoAccesos.fecha_venc_poliza.isnot(None))
        .all()
    ):
        if not a.medico or not a.fecha_venc_poliza:
            continue
        dias = (a.fecha_venc_poliza - hoy_d).days
        if dias <= 30:
            tipo = "HSM" if a.medico.tipo_listado == "cuerpo_medico" else "FSFB"
            alertas_30_enrich.append([
                tipo, a.medico.nombre_medico, "Póliza RC",
                a.fecha_venc_poliza.strftime("%d/%m/%Y"), str(dias),
            ])

    alertas_30_enrich.sort(key=lambda x: int(x[4]))
    alertas_count = len(set(a[1] for a in alertas_30_enrich))

    # ── FORMATO XLSX ──
    if formato == "xlsx":
        resumen_data = {
            "Indicador": [
                "Total Médicos", "Activos", "Finalizados", "Renuncias",
                "Inactivos", "Cuerpo Médico HSM", "FSFB Externos",
                "Alertas Venc. (≤30d)",
            ],
            "Valor": [
                total, activos, finalizados, renuncias,
                inactivos, hsm_total, fsfb_total, alertas_count,
            ],
        }
        df_resumen = pd.DataFrame(resumen_data)

        df_alertas = pd.DataFrame(
            alertas_30_enrich,
            columns=["Tipo", "Nombre", "Documento", "Fecha Vencimiento", "Días Restantes"],
        )

        return _styled_xlsx(
            {"Resumen Ejecutivo": df_resumen, "Alertas ≤30 días": df_alertas},
            "informe_ejecutivo.xlsx",
        )

    # ── FORMATO PDF (default) ──
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=landscape(A4),
        rightMargin=1.8 * cm, leftMargin=1.8 * cm,
        topMargin=2.2 * cm, bottomMargin=2.2 * cm,
    )
    W = landscape(A4)[0] - 3.6 * cm

    story = []

    # ── TÍTULO ──
    story.append(_pdf_section_title("INFORME EJECUTIVO — CUERPO MÉDICO", size=16, color=PDF_NAVY))
    story.append(_pdf_body_text(
        f"Hospital Serena del Mar  ·  Generado: {hoy.strftime('%d de %B de %Y').title()}", size=9
    ))
    story.append(Spacer(1, 0.4 * cm))
    story.append(HRFlowable(width="100%", thickness=2, color=PDF_JADE, spaceAfter=10))

    # ── TABLA 1: RESUMEN GENERAL ──
    story.append(_pdf_section_title("1. Resumen General", size=12, color=PDF_BLUE))
    story.append(Spacer(1, 0.25 * cm))

    resumen_data = [
        [_pdf_cell("INDICADOR", bold=True, size=9, color="FFFFFF", align="CENTER"),
         _pdf_cell("VALOR",     bold=True, size=9, color="FFFFFF", align="CENTER"),
         _pdf_cell("INDICADOR", bold=True, size=9, color="FFFFFF", align="CENTER"),
         _pdf_cell("VALOR",     bold=True, size=9, color="FFFFFF", align="CENTER")],
        [_pdf_cell("Total Médicos"),    _pdf_cell(str(total),       bold=True, color=BRAND_NAVY),
         _pdf_cell("Cuerpo Médico HSM"), _pdf_cell(str(hsm_total), bold=True, color=BRAND_NAVY)],
        [_pdf_cell("Activos"),          _pdf_cell(str(activos),     bold=True, color=BRAND_JADE),
         _pdf_cell("FSFB Externos"),    _pdf_cell(str(fsfb_total),  bold=True, color=BRAND_NAVY)],
        [_pdf_cell("Finalizados"),      _pdf_cell(str(finalizados), bold=True, color="#4F46E5"),
         _pdf_cell("Alertas Venc. (≤30d)"), _pdf_cell(str(alertas_count), bold=True,
                                              color=BRAND_RED if alertas_count > 0 else BRAND_JADE)],
        [_pdf_cell("Renuncias"),        _pdf_cell(str(renuncias),   bold=True, color=BRAND_AMBER),
         _pdf_cell("Inactivos"),        _pdf_cell(str(inactivos),   bold=True, color=BRAND_TEXT_MUTED)],
    ]

    col_w = W / 4
    resumen_table = Table(resumen_data, colWidths=[col_w * 1.6, col_w * 0.9, col_w * 1.6, col_w * 0.9])
    resumen_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), PDF_BLUE),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("GRID",          (0, 0), (-1, -1), 0.5, PDF_MGREY),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PDF_LGREY]),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("LINEABOVE",     (0, 0), (-1, 0), 1, PDF_BLUE),
    ]))
    story.append(resumen_table)
    story.append(Spacer(1, 0.6 * cm))

    # ── TABLA 2: DOCUMENTOS POR VENCER ──
    story.append(_pdf_section_title(
        "2. Médicos con Documentos por Vencer (≤ 30 días)", size=12, color=PDF_BLUE
    ))
    story.append(Spacer(1, 0.25 * cm))

    if alertas_30_enrich:
        header_row = [
            _pdf_cell("TIPO",              bold=True, size=8, color="FFFFFF", align="CENTER"),
            _pdf_cell("NOMBRE",            bold=True, size=8, color="FFFFFF", align="CENTER"),
            _pdf_cell("DOCUMENTO",         bold=True, size=8, color="FFFFFF", align="CENTER"),
            _pdf_cell("FECHA VENCIMIENTO", bold=True, size=8, color="FFFFFF", align="CENTER"),
            _pdf_cell("DÍAS RESTANTES",    bold=True, size=8, color="FFFFFF", align="CENTER"),
        ]
        rows_pdf = [header_row]
        for row in alertas_30_enrich:
            dias_int = int(row[4])
            dias_color = BRAND_RED if dias_int < 0 else (BRAND_AMBER if dias_int <= 7 else BRAND_TEXT_DARK)
            badge_text = "VENCIDO" if dias_int < 0 else f"{dias_int}d"
            rows_pdf.append([
                _pdf_cell(row[0], bold=True,
                          color=BRAND_JADE if row[0] == "HSM" else "#1A4ED7", align="CENTER"),
                _pdf_cell(row[1], size=8),
                _pdf_cell(row[2], size=8),
                _pdf_cell(row[3], size=8, align="CENTER"),
                _pdf_cell(badge_text, bold=True, size=8, color=dias_color, align="CENTER"),
            ])

        cw = [W * 0.07, W * 0.30, W * 0.27, W * 0.18, W * 0.15]
        alertas_table = Table(rows_pdf, colWidths=cw, repeatRows=1)
        alertas_table.setStyle(_pdf_standard_table_style(header_bg=PDF_NAVY))
        story.append(alertas_table)
        story.append(Spacer(1, 0.3 * cm))
        story.append(_pdf_body_text(
            f"Total: {len(alertas_30_enrich)} alerta(s) activa(s) · "
            f"{sum(1 for r in alertas_30_enrich if int(r[4]) < 0)} ya vencida(s)", size=8
        ))
    else:
        story.append(_pdf_body_text(
            "No hay documentos por vencer en los próximos 30 días.", size=9
        ))

    doc.build(story, onFirstPage=_pdf_header_footer, onLaterPages=_pdf_header_footer)
    output.seek(0)

    resp_headers = {"Content-Disposition": 'attachment; filename="informe_ejecutivo.pdf"'}
    return StreamingResponse(
        iter([output.getvalue()]),
        headers=resp_headers,
        media_type="application/pdf",
    )
